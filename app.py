import streamlit as st
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging

# Set up logging for debugging (visible in console, not UI)
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

st.set_page_config(page_title="Excel Files Consolidator", layout="wide")
st.title("📊 Excel Files Consolidator (Data Table Sheet Only)")

# --- INPUT FOLDER ---
input_folder = st.text_input(
    "📂 Enter the INPUT folder path:",
    r"C:\Users\qimatest\OneDrive - QIMA\Accounting & Finance Team - People Planning - BO Review"
)

files = []
if input_folder and os.path.exists(input_folder):
    for root, dirs, filenames in os.walk(input_folder):
        for f in filenames:
            if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$"):
                files.append(os.path.join(root, f))

# --- SELECT ALL OPTION ---
nice_file_list = [os.path.relpath(f, input_folder) for f in files]
select_all = st.checkbox("🔘 Select All Files")
selected_display = st.multiselect(
    "✅ Select Excel files to consolidate (from all subfolders):",
    options=nice_file_list,
    default=nice_file_list if select_all else []
)
selected_files = [os.path.join(input_folder, f) for f in selected_display]

# --- OUTPUT FOLDER ---
output_folder = st.text_input(
    "💾 Enter the OUTPUT folder path:",
    r"C:\Users\qimatest\OneDrive - QIMA\Accounting & Finance Team - People Planning - BO Review\CP_Budget_Consolidated"
)

# --- PARAMETERS ---
HEADER_ROWS = 12  # Headers up to row 12
TARGET_SHEET = "Data Table"  # Only process this sheet

# --- CONSOLIDATE ---
if st.button("🚀 Consolidate Files"):
    if not selected_files:
        st.error("Please select at least one file.")
    else:
        try:
            # Validate and create output folder
            output_folder = os.path.abspath(output_folder)
            if not os.path.exists(output_folder):
                logger.info(f"Creating output folder: {output_folder}")
                os.makedirs(output_folder, exist_ok=True)
            st.info(f"Output folder: {output_folder}")

            # Test write permissions
            output_path = os.path.join(output_folder, "Consolidated_File.xlsx")
            output_path = os.path.abspath(output_path)
            try:
                with open(os.path.join(output_folder, "test_write.txt"), "w") as f:
                    f.write("Test")
                os.remove(os.path.join(output_folder, "test_write.txt"))
            except Exception as e:
                st.error(f"❌ Cannot write to output folder: {e}")
                raise

            # Initialize progress bars and status text
            file_progress_bar = st.progress(0)
            status_text = st.empty()
            total_files = len(selected_files)

            # Load template workbook for headers (first file with formulas)
            template_file = selected_files[0]
            logger.info(f"Loading template file: {template_file}")
            wb_master = load_workbook(template_file, keep_vba=False, data_only=False)
            if TARGET_SHEET not in wb_master.sheetnames:
                st.error(f"❌ Template file does not contain '{TARGET_SHEET}' sheet")
                raise ValueError(f"Template file does not contain '{TARGET_SHEET}' sheet")
            
            # Remove all sheets except TARGET_SHEET
            for sheet_name in wb_master.sheetnames:
                if sheet_name != TARGET_SHEET:
                    wb_master.remove(wb_master[sheet_name])
            ws_master = wb_master[TARGET_SHEET]

            # Get template column headers and max column
            header_row = next(ws_master.iter_rows(min_row=HEADER_ROWS, max_row=HEADER_ROWS, max_col=ws_master.max_column))
            headers = [cell.value for cell in header_row]
            max_col = len(headers)
            logger.info(f"Template sheet: {ws_master.max_row} rows, {max_col} columns")

            # Clear data rows from template (row 13 onwards) and copy without formatting
            if ws_master.max_row > HEADER_ROWS:
                logger.info("Clearing data rows from template and copying values only")
                data_rows = []
                for row in ws_master.iter_rows(min_row=HEADER_ROWS + 1, max_row=ws_master.max_row, max_col=max_col):
                    row_data = [round(cell.value) if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool) else cell.value
                                for cell in row]
                    data_rows.append(row_data)
                ws_master.delete_rows(HEADER_ROWS + 1, ws_master.max_row - HEADER_ROWS)
                # Write data rows back
                for row_data in data_rows:
                    ws_master.append(row_data[:max_col])

            # Process each file (including reprocessing the first file for consistency)
            for idx, file_path in enumerate(selected_files, start=1):
                # Update file progress
                file_percentage = idx / total_files
                file_progress_bar.progress(file_percentage)
                status_text.text(f"Processing: {os.path.basename(file_path)} ({idx}/{total_files}, {file_percentage:.0%})")
                logger.info(f"Processing file: {file_path}")

                # Load source workbook
                try:
                    wb_src = load_workbook(file_path, read_only=True, data_only=True)
                    if TARGET_SHEET not in wb_src.sheetnames:
                        logger.warning(f"Skipping {file_path}: '{TARGET_SHEET}' sheet not found")
                        wb_src.close()
                        continue
                    ws_src = wb_src[TARGET_SHEET]
                    src_max_row = ws_src.max_row
                    src_max_col = min(ws_src.max_column, max_col)
                    logger.info(f"Source sheet: {src_max_row} rows, {src_max_col} columns")

                    # Read all data rows into memory
                    data_rows = []
                    for row in ws_src.iter_rows(min_row=HEADER_ROWS + 1, max_row=src_max_row, max_col=src_max_col):
                        row_data = [round(cell.value) if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool) else cell.value
                                    for cell in row]
                        data_rows.append(row_data)

                    # Write all data rows to master sheet at once
                    for row_data in data_rows:
                        ws_master.append(row_data[:max_col])

                    wb_src.close()
                except Exception as e:
                    logger.warning(f"Error processing {file_path}: {e}")
                    continue

            # Save as final .xlsx
            logger.info(f"Saving consolidated file to: {output_path}")
            wb_master.save(output_path)

            # Verify file was created
            if os.path.exists(output_path):
                st.success(f"✅ Consolidation complete! File saved at: {output_path}")
                with open(output_path, "rb") as f:
                    st.download_button(
                        label="⬇ Download Consolidated File",
                        data=f,
                        file_name="Consolidated_File.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.error(f"❌ File not found at: {output_path}. Check folder permissions or disk space.")

        except Exception as e:
            logger.error(f"Error during consolidation: {e}")
            status_text.error(f"❌ Error during consolidation: {e}")

        finally:
            file_progress_bar.empty()
            status_text.empty()
